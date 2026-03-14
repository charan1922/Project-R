"""
JINDALSTEL Deep Dive — Pull every data point, export to Excel
Target: R Factor = 3.3590890034909253 (March 13, 2026)
"""
import requests, csv, io, time, json, math
from datetime import datetime, timedelta
from collections import defaultdict

# ============ CREDENTIALS ============
ACCESS_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpc3MiOiJkaGFuIiwicGFydG5lcklkIjoiIiwiZXhwIjoxNzczNTEwMzgzLCJpYXQiOjE3NzM0MjM5ODMsInRva2VuQ29uc3VtZXJUeXBlIjoiU0VMRiIsIndlYmhvb2tVcmwiOiIiLCJkaGFuQ2xpZW50SWQiOiIxMTA2ODAwMzk0In0.Q8o8vpzrrAMO7SeRPIVkUh47GuKtxHmvQUn5dPQo1ooWYHtm77EaTtVlBfaVVbEhKlbb7MvDsXakNnuCHf3Azw"
CLIENT_ID = "1106800394"
BASE_URL = "https://api.dhan.co/v2"
HEADERS = {"access-token": ACCESS_TOKEN, "Content-Type": "application/json"}
HEADERS_OC = {"access-token": ACCESS_TOKEN, "client-id": CLIENT_ID, "Content-Type": "application/json"}

TARGET_DATE = "2026-03-13"
ACTUAL_RFACTOR = 3.3590890034909253
SYMBOL = "JINDALSTEL"

# Also pull a LOW R-factor stock for comparison
COMPARE_SYMBOL = "INDIANB"
COMPARE_RFACTOR = 1.578396175596205

api_calls = 0

def api_call(url, payload, headers=HEADERS, rate=0.3):
    global api_calls
    resp = requests.post(f"{BASE_URL}{url}", headers=headers, json=payload)
    api_calls += 1
    time.sleep(rate)
    if resp.status_code == 200:
        return resp.json()
    print(f"  API error {resp.status_code}: {url} -> {resp.text[:200]}")
    return None

def api_historical(security_id, exchange_segment, instrument, from_date, to_date, oi=True):
    data = api_call("/charts/historical", {
        "securityId": str(security_id), "exchangeSegment": exchange_segment,
        "instrument": instrument, "oi": oi, "fromDate": from_date, "toDate": to_date
    })
    if data and 'open' in data and len(data['open']) > 0:
        return data
    return None

def api_option_chain(equity_sec_id, expiry_date):
    data = api_call("/optionchain", {
        "UnderlyingScrip": int(equity_sec_id), "UnderlyingSeg": "NSE_FNO", "Expiry": expiry_date
    }, headers=HEADERS_OC, rate=3)
    if data:
        d = data.get('data', data)
        if 'oc' in d:
            return d
    return None

def api_expiry_list(equity_sec_id):
    data = api_call("/optionchain/expirylist", {
        "UnderlyingScrip": int(equity_sec_id), "UnderlyingSeg": "NSE_FNO"
    }, headers=HEADERS_OC, rate=3)
    if data:
        d = data.get('data', data)
        if isinstance(d, list):
            return d
    return []


def download_scrip_master():
    print("Downloading scrip master...")
    resp = requests.get("https://images.dhan.co/api-data/api-scrip-master.csv")
    reader = csv.DictReader(io.StringIO(resp.text))
    return list(reader)


def build_maps_for_symbol(rows, symbol):
    """Build equity, futures, and ALL options contracts for a symbol"""
    target = datetime.strptime(TARGET_DATE, '%Y-%m-%d').date()
    equity = None
    futures = None
    options = []

    for row in rows:
        try:
            if row.get('SEM_EXM_EXCH_ID') != 'NSE':
                continue
            sym = row.get('SEM_TRADING_SYMBOL', '')
            inst = row.get('SEM_INSTRUMENT_NAME', '')
            sec_id = row.get('SEM_SMST_SECURITY_ID', '')
            base = sym.split('-')[0]
            if base != symbol:
                continue

            if inst == 'EQUITY':
                equity = {'security_id': sec_id, 'symbol': sym}

            elif inst == 'FUTSTK':
                expiry_str = row.get('SEM_EXPIRY_DATE', '')
                if expiry_str:
                    expiry = datetime.strptime(expiry_str.split(' ')[0], '%Y-%m-%d').date()
                    if expiry >= target:
                        if futures is None or expiry < futures['expiry']:
                            futures = {'security_id': sec_id, 'expiry': expiry, 'symbol': sym}

            elif inst == 'OPTSTK':
                expiry_str = row.get('SEM_EXPIRY_DATE', '')
                if expiry_str and '14:30' in expiry_str:  # Monthly only
                    expiry = datetime.strptime(expiry_str.split(' ')[0], '%Y-%m-%d').date()
                    if expiry >= target:
                        strike = float(row.get('SEM_STRIKE_PRICE', 0))
                        opt_type = row.get('SEM_OPTION_TYPE', '')
                        options.append({
                            'security_id': sec_id, 'strike': strike,
                            'opt_type': opt_type, 'expiry': expiry, 'symbol': sym
                        })
        except Exception:
            continue

    return equity, futures, options


def parse_historical_to_rows(data):
    """Convert API response to list of dicts with date, open, high, low, close, volume, oi"""
    rows = []
    ts_arr = data.get('timestamp', [])
    for i, ts in enumerate(ts_arr):
        dt = datetime.fromtimestamp(ts)
        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'open': data['open'][i],
            'high': data['high'][i],
            'low': data['low'][i],
            'close': data['close'][i],
            'volume': data.get('volume', [0]*len(ts_arr))[i],
        }
        if 'open_interest' in data and i < len(data['open_interest']):
            row['oi'] = data['open_interest'][i]
        rows.append(row)
    return rows


def collect_single_stock(rows, symbol, actual_r):
    """Collect ALL data for a single stock"""
    print(f"\n{'='*70}")
    print(f"  {symbol}  |  Actual R Factor = {actual_r:.4f}")
    print(f"{'='*70}")

    equity, futures, options = build_maps_for_symbol(rows, symbol)
    print(f"  Equity: {equity is not None} | Futures: {futures is not None} | Options: {len(options)} contracts")

    target_dt = datetime.strptime(TARGET_DATE, '%Y-%m-%d')
    from_date = (target_dt - timedelta(days=45)).strftime('%Y-%m-%d')

    result = {'symbol': symbol, 'actual_rfactor': actual_r}

    # ---- 1. EQUITY historical ----
    eq_rows = []
    if equity:
        data = api_historical(equity['security_id'], "NSE_EQ", "EQUITY", from_date, TARGET_DATE, oi=False)
        if data:
            eq_rows = parse_historical_to_rows(data)
            print(f"  Equity: {len(eq_rows)} days, LTP={eq_rows[-1]['close']}")
    result['equity_daily'] = eq_rows

    # ---- 2. FUTURES historical ----
    fut_rows = []
    if futures:
        data = api_historical(futures['security_id'], "NSE_FNO", "FUTSTK", from_date, TARGET_DATE)
        if data:
            fut_rows = parse_historical_to_rows(data)
            print(f"  Futures: {len(fut_rows)} days, OI={fut_rows[-1].get('oi', 0):,}")
    result['futures_daily'] = fut_rows

    # ---- 3. OPTIONS historical — ALL near-month strikes near ATM ----
    ltp = eq_rows[-1]['close'] if eq_rows else 0
    opt_data = {}  # {(strike, type): [daily_rows]}
    opt_daily_total = defaultdict(float)  # date -> total OI across all strikes

    if options and ltp > 0:
        # Filter to nearest monthly expiry
        nearest_expiry = min(set(o['expiry'] for o in options))
        near_opts = [o for o in options if o['expiry'] == nearest_expiry]
        all_strikes = sorted(set(o['strike'] for o in near_opts))
        atm_strike = min(all_strikes, key=lambda s: abs(s - ltp))
        atm_idx = all_strikes.index(atm_strike)

        # ATM ± 5 strikes (wider than before)
        lo = max(0, atm_idx - 5)
        hi = min(len(all_strikes), atm_idx + 6)
        target_strikes = all_strikes[lo:hi]
        print(f"  Options: ATM={atm_strike} (LTP={ltp}), pulling {len(target_strikes)} strikes: {target_strikes}")
        print(f"  Expiry: {nearest_expiry}")

        for strike in target_strikes:
            for opt_type in ['CE', 'PE']:
                contract = next((o for o in near_opts if o['strike'] == strike and o['opt_type'] == opt_type), None)
                if not contract:
                    continue
                data = api_historical(contract['security_id'], "NSE_FNO", "OPTSTK", from_date, TARGET_DATE)
                if data:
                    opt_rows = parse_historical_to_rows(data)
                    key = f"{int(strike)}{opt_type}"
                    opt_data[key] = opt_rows
                    for r in opt_rows:
                        opt_daily_total[r['date']] += r.get('oi', 0)
                    last_oi = opt_rows[-1].get('oi', 0) if opt_rows else 0
                    print(f"    {key}: {len(opt_rows)} days, last OI={last_oi:,}")

    result['options_contracts'] = {k: v for k, v in opt_data.items()}
    result['options_daily_total_oi'] = dict(opt_daily_total)
    result['ltp'] = ltp

    # ---- 4. Option Chain snapshot ----
    oc_data = None
    if equity:
        expiries = api_expiry_list(equity['security_id'])
        if expiries:
            oc_data = api_option_chain(equity['security_id'], expiries[0])
            if oc_data:
                oc = oc_data.get('oc', {})
                total_oi = sum(
                    (s.get('ce', {}).get('oi', 0) or 0) + (s.get('pe', {}).get('oi', 0) or 0)
                    for s in oc.values()
                )
                print(f"  Option Chain: {len(oc)} strikes, Total OI={total_oi:,}")
    result['option_chain'] = oc_data

    return result


def compute_metrics(result):
    """Compute all relative metrics from raw data"""
    metrics = {'symbol': result['symbol'], 'actual_rfactor': result['actual_rfactor']}

    # --- Equity metrics ---
    eq = result['equity_daily']
    if len(eq) >= 5:
        today = eq[-1]
        hist = eq[:-1][-20:]
        n = len(hist)

        today_vol = today['volume']
        today_turn = today['volume'] * today['close']
        avg_vol = sum(r['volume'] for r in hist) / n
        avg_turn = sum(r['volume'] * r['close'] for r in hist) / n

        metrics['eq_rel_volume'] = today_vol / avg_vol if avg_vol > 0 else 0
        metrics['eq_rel_turnover'] = today_turn / avg_turn if avg_turn > 0 else 0
        metrics['eq_today_volume'] = today_vol
        metrics['eq_avg_volume'] = avg_vol
        metrics['eq_today_turnover'] = today_turn
        metrics['eq_avg_turnover'] = avg_turn
        metrics['eq_ltp'] = today['close']

    # --- Futures metrics ---
    fut = result['futures_daily']
    if len(fut) >= 5:
        today = fut[-1]
        hist = fut[:-1][-20:]
        n = len(hist)

        today_oi = today.get('oi', 0)
        today_vol = today['volume']
        today_turn = today['volume'] * today['close']
        avg_oi = sum(r.get('oi', 0) for r in hist) / n
        avg_vol = sum(r['volume'] for r in hist) / n
        avg_turn = sum(r['volume'] * r['close'] for r in hist) / n

        oi_changes = [abs(hist[i].get('oi', 0) - hist[i-1].get('oi', 0)) for i in range(1, n)]
        avg_oi_change = sum(oi_changes) / len(oi_changes) if oi_changes else 1
        today_oi_change = abs(today_oi - hist[-1].get('oi', 0))

        metrics['fut_rel_oi'] = today_oi / avg_oi if avg_oi > 0 else 0
        metrics['fut_rel_oi_change'] = today_oi_change / avg_oi_change if avg_oi_change > 0 else 0
        metrics['fut_rel_volume'] = today_vol / avg_vol if avg_vol > 0 else 0
        metrics['fut_rel_turnover'] = today_turn / avg_turn if avg_turn > 0 else 0
        metrics['fut_today_oi'] = today_oi
        metrics['fut_avg_oi'] = avg_oi
        metrics['fut_today_volume'] = today_vol
        metrics['fut_avg_volume'] = avg_vol
        metrics['fut_today_turnover'] = today_turn
        metrics['fut_avg_turnover'] = avg_turn

    # --- Options OI metrics (summed across all strikes) ---
    opt_oi = result['options_daily_total_oi']
    if len(opt_oi) >= 5:
        dates = sorted(opt_oi.keys())
        values = [opt_oi[d] for d in dates]
        today_oi = values[-1]
        hist_oi = values[:-1][-20:]
        avg_oi = sum(hist_oi) / len(hist_oi)
        oi_changes = [abs(hist_oi[i] - hist_oi[i-1]) for i in range(1, len(hist_oi))]
        avg_change = sum(oi_changes) / len(oi_changes) if oi_changes else 1
        today_change = abs(today_oi - hist_oi[-1])

        metrics['opt_rel_oi'] = today_oi / avg_oi if avg_oi > 0 else 0
        metrics['opt_rel_oi_change'] = today_change / avg_change if avg_change > 0 else 0
        metrics['opt_today_oi'] = today_oi
        metrics['opt_avg_oi'] = avg_oi

    # --- Option Chain snapshot metrics ---
    oc_data = result.get('option_chain')
    if oc_data:
        oc = oc_data.get('oc', {})
        ltp = oc_data.get('last_price', result.get('ltp', 0))
        total_oi = 0
        total_prev_oi = 0
        total_vol = 0
        total_prev_vol = 0
        weighted_spread = 0
        spread_weight = 0

        for strike_key, strike_data in oc.items():
            strike = float(strike_key)
            dist = abs(strike - ltp) / ltp if ltp > 0 else 1
            weight = max(0, 1 - dist * 5)

            for t in ['ce', 'pe']:
                if t in strike_data:
                    d = strike_data[t]
                    oi = d.get('oi', 0) or 0
                    prev_oi = d.get('previous_oi', 0) or 0
                    vol = d.get('volume', 0) or 0
                    prev_vol = d.get('previous_volume', 0) or 0
                    total_oi += oi
                    total_prev_oi += prev_oi
                    total_vol += vol
                    total_prev_vol += prev_vol

                    bid = d.get('top_bid_price', 0) or 0
                    ask = d.get('top_ask_price', 0) or 0
                    if bid > 0 and ask > 0:
                        spread_pct = (ask - bid) / ((ask + bid) / 2)
                        weighted_spread += spread_pct * weight
                        spread_weight += weight

        metrics['oc_total_oi'] = total_oi
        metrics['oc_prev_oi'] = total_prev_oi
        metrics['oc_oi_change'] = total_oi - total_prev_oi
        metrics['oc_oi_change_pct'] = (total_oi - total_prev_oi) / total_prev_oi if total_prev_oi > 0 else 0
        metrics['oc_total_volume'] = total_vol
        metrics['oc_vol_ratio'] = total_vol / total_prev_vol if total_prev_vol > 0 else 0
        metrics['oc_avg_spread'] = weighted_spread / spread_weight if spread_weight > 0 else 0

    return metrics


def export_to_excel(primary, compare=None):
    """Export everything to Excel with multiple sheets"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call(['pip3', 'install', 'openpyxl'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = Workbook()
    header_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def write_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font

    # ===== Sheet 1: Summary Metrics =====
    ws = wb.active
    ws.title = "Summary"
    p_metrics = compute_metrics(primary)
    c_metrics = compute_metrics(compare) if compare else None

    write_header(ws, ["Metric", SYMBOL, COMPARE_SYMBOL if compare else "", "Ratio (High/Low)"])
    row = 2
    metric_keys = [k for k in p_metrics if k not in ('symbol', 'actual_rfactor')]
    ws.cell(row=row, column=1, value="Actual R Factor")
    ws.cell(row=row, column=2, value=p_metrics['actual_rfactor'])
    if c_metrics:
        ws.cell(row=row, column=3, value=c_metrics['actual_rfactor'])
    row += 1

    for k in sorted(metric_keys):
        ws.cell(row=row, column=1, value=k)
        pv = p_metrics.get(k, 0)
        ws.cell(row=row, column=2, value=round(pv, 6) if isinstance(pv, float) else pv)
        if c_metrics:
            cv = c_metrics.get(k, 0)
            ws.cell(row=row, column=3, value=round(cv, 6) if isinstance(cv, float) else cv)
            if cv and pv and cv != 0:
                ws.cell(row=row, column=4, value=round(pv / cv, 4))
        row += 1

    # ===== Sheet 2: Equity Daily =====
    ws2 = wb.create_sheet("Equity Daily")
    eq_headers = ["Date", f"{SYMBOL}_Close", f"{SYMBOL}_Volume", f"{SYMBOL}_Turnover"]
    if compare:
        eq_headers += [f"{COMPARE_SYMBOL}_Close", f"{COMPARE_SYMBOL}_Volume", f"{COMPARE_SYMBOL}_Turnover"]
    write_header(ws2, eq_headers)

    # Merge dates
    p_eq = {r['date']: r for r in primary['equity_daily']}
    c_eq = {r['date']: r for r in (compare['equity_daily'] if compare else [])}
    all_dates = sorted(set(list(p_eq.keys()) + list(c_eq.keys())))

    for i, dt in enumerate(all_dates, 2):
        ws2.cell(row=i, column=1, value=dt)
        if dt in p_eq:
            r = p_eq[dt]
            ws2.cell(row=i, column=2, value=r['close'])
            ws2.cell(row=i, column=3, value=r['volume'])
            ws2.cell(row=i, column=4, value=round(r['close'] * r['volume']))
        if compare and dt in c_eq:
            r = c_eq[dt]
            ws2.cell(row=i, column=5, value=r['close'])
            ws2.cell(row=i, column=6, value=r['volume'])
            ws2.cell(row=i, column=7, value=round(r['close'] * r['volume']))

    # ===== Sheet 3: Futures Daily =====
    ws3 = wb.create_sheet("Futures Daily")
    fut_headers = ["Date", f"{SYMBOL}_Close", f"{SYMBOL}_Volume", f"{SYMBOL}_OI", f"{SYMBOL}_Turnover"]
    if compare:
        fut_headers += [f"{COMPARE_SYMBOL}_Close", f"{COMPARE_SYMBOL}_Volume", f"{COMPARE_SYMBOL}_OI", f"{COMPARE_SYMBOL}_Turnover"]
    write_header(ws3, fut_headers)

    p_fut = {r['date']: r for r in primary['futures_daily']}
    c_fut = {r['date']: r for r in (compare['futures_daily'] if compare else [])}
    all_dates = sorted(set(list(p_fut.keys()) + list(c_fut.keys())))

    for i, dt in enumerate(all_dates, 2):
        ws3.cell(row=i, column=1, value=dt)
        if dt in p_fut:
            r = p_fut[dt]
            ws3.cell(row=i, column=2, value=r['close'])
            ws3.cell(row=i, column=3, value=r['volume'])
            ws3.cell(row=i, column=4, value=r.get('oi', 0))
            ws3.cell(row=i, column=5, value=round(r['close'] * r['volume']))
        if compare and dt in c_fut:
            r = c_fut[dt]
            ws3.cell(row=i, column=6, value=r['close'])
            ws3.cell(row=i, column=7, value=r['volume'])
            ws3.cell(row=i, column=8, value=r.get('oi', 0))
            ws3.cell(row=i, column=9, value=round(r['close'] * r['volume']))

    # ===== Sheet 4: Options OI by Strike (JINDALSTEL) =====
    ws4 = wb.create_sheet(f"{SYMBOL} Options OI")
    contracts = primary.get('options_contracts', {})
    if contracts:
        # Headers: Date, then each contract
        contract_names = sorted(contracts.keys())
        write_header(ws4, ["Date"] + contract_names + ["TOTAL_OI"])

        # Collect all dates
        all_opt_dates = set()
        for cname, crows in contracts.items():
            for r in crows:
                all_opt_dates.add(r['date'])
        all_opt_dates = sorted(all_opt_dates)

        # Build lookup
        lookup = {}
        for cname, crows in contracts.items():
            for r in crows:
                lookup[(r['date'], cname)] = r.get('oi', 0)

        for i, dt in enumerate(all_opt_dates, 2):
            ws4.cell(row=i, column=1, value=dt)
            total = 0
            for j, cname in enumerate(contract_names, 2):
                oi = lookup.get((dt, cname), 0)
                ws4.cell(row=i, column=j, value=oi)
                total += oi
            ws4.cell(row=i, column=len(contract_names) + 2, value=total)

    # ===== Sheet 5: Options OI by Strike (Compare) =====
    if compare and compare.get('options_contracts'):
        ws5 = wb.create_sheet(f"{COMPARE_SYMBOL} Options OI")
        contracts = compare['options_contracts']
        contract_names = sorted(contracts.keys())
        write_header(ws5, ["Date"] + contract_names + ["TOTAL_OI"])

        all_opt_dates = set()
        for cname, crows in contracts.items():
            for r in crows:
                all_opt_dates.add(r['date'])
        all_opt_dates = sorted(all_opt_dates)

        lookup = {}
        for cname, crows in contracts.items():
            for r in crows:
                lookup[(r['date'], cname)] = r.get('oi', 0)

        for i, dt in enumerate(all_opt_dates, 2):
            ws5.cell(row=i, column=1, value=dt)
            total = 0
            for j, cname in enumerate(contract_names, 2):
                oi = lookup.get((dt, cname), 0)
                ws5.cell(row=i, column=j, value=oi)
                total += oi
            ws5.cell(row=i, column=len(contract_names) + 2, value=total)

    # ===== Sheet 6: Option Chain Snapshot (JINDALSTEL) =====
    oc_data = primary.get('option_chain')
    if oc_data:
        ws6 = wb.create_sheet(f"{SYMBOL} OptionChain")
        write_header(ws6, [
            "Strike", "CE_OI", "CE_PrevOI", "CE_OI_Chg", "CE_Volume", "CE_PrevVol",
            "CE_Bid", "CE_Ask", "CE_Spread%", "CE_IV",
            "PE_OI", "PE_PrevOI", "PE_OI_Chg", "PE_Volume", "PE_PrevVol",
            "PE_Bid", "PE_Ask", "PE_Spread%", "PE_IV",
            "Total_OI", "Total_OI_Chg"
        ])
        oc = oc_data.get('oc', {})
        ltp = oc_data.get('last_price', primary.get('ltp', 0))
        row = 2
        for strike_key in sorted(oc.keys(), key=lambda x: float(x)):
            s = oc[strike_key]
            strike = float(strike_key)
            ws6.cell(row=row, column=1, value=strike)

            total_oi = 0
            total_chg = 0
            col = 2
            for t in ['ce', 'pe']:
                d = s.get(t, {})
                oi = d.get('oi', 0) or 0
                prev_oi = d.get('previous_oi', 0) or 0
                vol = d.get('volume', 0) or 0
                prev_vol = d.get('previous_volume', 0) or 0
                bid = d.get('top_bid_price', 0) or 0
                ask = d.get('top_ask_price', 0) or 0
                iv = d.get('implied_volatility', 0) or 0
                spread = (ask - bid) / ((ask + bid) / 2) * 100 if bid > 0 and ask > 0 else 0

                ws6.cell(row=row, column=col, value=oi); col += 1
                ws6.cell(row=row, column=col, value=prev_oi); col += 1
                ws6.cell(row=row, column=col, value=oi - prev_oi); col += 1
                ws6.cell(row=row, column=col, value=vol); col += 1
                ws6.cell(row=row, column=col, value=prev_vol); col += 1
                ws6.cell(row=row, column=col, value=bid); col += 1
                ws6.cell(row=row, column=col, value=ask); col += 1
                ws6.cell(row=row, column=col, value=round(spread, 2)); col += 1
                ws6.cell(row=row, column=col, value=round(iv, 2)); col += 1

                total_oi += oi
                total_chg += (oi - prev_oi)

            ws6.cell(row=row, column=col, value=total_oi); col += 1
            ws6.cell(row=row, column=col, value=total_chg)

            # Highlight ATM
            if abs(strike - ltp) < (ltp * 0.02):
                for c in range(1, col + 1):
                    ws6.cell(row=row, column=c).fill = green_fill
            row += 1

    # ===== Sheet 7: Relative Metrics Timeline =====
    ws7 = wb.create_sheet("Relative Metrics")
    write_header(ws7, [
        "Date",
        f"{SYMBOL}_EqRelVol", f"{SYMBOL}_EqRelTurn",
        f"{SYMBOL}_FutRelOI", f"{SYMBOL}_FutRelVol", f"{SYMBOL}_FutRelTurn",
        f"{SYMBOL}_OptRelOI",
    ])
    if compare:
        # Add compare columns starting at col 8
        for h in [f"{COMPARE_SYMBOL}_EqRelVol", f"{COMPARE_SYMBOL}_EqRelTurn",
                   f"{COMPARE_SYMBOL}_FutRelOI", f"{COMPARE_SYMBOL}_FutRelVol",
                   f"{COMPARE_SYMBOL}_FutRelTurn", f"{COMPARE_SYMBOL}_OptRelOI"]:
            ws7.cell(row=1, column=ws7.max_column + 1, value=h).font = header_font

    # Compute rolling relative metrics for each day
    def compute_rolling(daily_rows, opt_daily_total, start_col, ws_ref):
        for day_idx in range(5, len(daily_rows)):
            today = daily_rows[day_idx]
            hist = daily_rows[max(0, day_idx-20):day_idx]
            n = len(hist)
            dt = today['date']

            avg_vol = sum(r['volume'] for r in hist) / n
            avg_turn = sum(r['volume'] * r['close'] for r in hist) / n
            today_vol = today['volume']
            today_turn = today['volume'] * today['close']

            row_num = day_idx - 4 + 1  # offset for header
            ws_ref.cell(row=row_num, column=1, value=dt)
            ws_ref.cell(row=row_num, column=start_col, value=round(today_vol / avg_vol, 4) if avg_vol > 0 else 0)
            ws_ref.cell(row=row_num, column=start_col + 1, value=round(today_turn / avg_turn, 4) if avg_turn > 0 else 0)

    # For now, compute the final-day snapshot per stock in the summary sheet
    # The rolling is complex with options, keep the sheet simple

    fname = f"derive-r/{SYMBOL}_deep_dive.xlsx"
    wb.save(fname)
    print(f"\nSaved Excel: {fname}")
    return fname


def main():
    rows = download_scrip_master()

    print(f"\nCollecting {SYMBOL} (R={ACTUAL_RFACTOR:.4f})...")
    primary = collect_single_stock(rows, SYMBOL, ACTUAL_RFACTOR)

    print(f"\nCollecting {COMPARE_SYMBOL} (R={COMPARE_RFACTOR:.4f}) for comparison...")
    compare = collect_single_stock(rows, COMPARE_SYMBOL, COMPARE_RFACTOR)

    # Save raw JSON
    raw = {'primary': primary, 'compare': compare}
    with open(f'derive-r/{SYMBOL}_deep_data.json', 'w') as f:
        json.dump(raw, f, indent=2, default=str)
    print(f"\nSaved raw data: derive-r/{SYMBOL}_deep_data.json")

    # Print key metrics side by side
    p_m = compute_metrics(primary)
    c_m = compute_metrics(compare)

    print(f"\n{'='*70}")
    print(f"{'METRIC':<25} {SYMBOL:>15} {COMPARE_SYMBOL:>15} {'RATIO':>10}")
    print(f"{'':25} {'R='+str(round(ACTUAL_RFACTOR,2)):>15} {'R='+str(round(COMPARE_RFACTOR,2)):>15}")
    print(f"{'='*70}")

    for k in sorted(p_m.keys()):
        if k in ('symbol', 'actual_rfactor'):
            continue
        pv = p_m.get(k, 0)
        cv = c_m.get(k, 0)
        ratio = pv / cv if cv and cv != 0 else 0
        print(f"  {k:<23} {pv:>15.4f} {cv:>15.4f} {ratio:>10.3f}")

    # Export to Excel
    export_to_excel(primary, compare)

    print(f"\nTotal API calls: {api_calls}")


if __name__ == "__main__":
    main()
